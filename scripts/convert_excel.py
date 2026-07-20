#!/usr/bin/env python3
"""
公式サイト配布の Excel を消費アトラスのデータ形式に変換するスクリプト。
(API 提供のない「貨物地域流動調査」と「県民経済計算」向け)

事前準備:
  pip install openpyxl
  ※ 拡張子 .xls の場合は Excel で .xlsx に保存し直してから使ってください。

■ 物流 (貨物地域流動調査 → data/flow_data.json)
  入手先: e-Stat「貨物地域流動調査」 https://www.e-stat.go.jp/stat-search/files?toukei=00600460
          → 対象年度 → 「統計表2 府県相互間輸送トン数表 (総貨物及び9品目分類)」の Excel
  変換:   python convert_excel.py matrix 表2.xlsx --year 2023
          # 行=発地 / 列=着地 として 47×47 の表を自動検出します
          # 向きが逆の表の場合は --swap、シート指定は --sheet、単位表記は --unit

■ 域際収支 (県民経済計算 → data/pref_data.json の指標 netexp)
  入手先: 内閣府「県民経済計算」統計表
          https://www.esri.cao.go.jp/jp/sna/data/data_list/kenmin/files/files_kenmin.html
          → 「財貨・サービスの移出入 (純)・統計上の不突合」を含む Excel (単位: 百万円)
  変換:   python convert_excel.py kenmin 移出入.xlsx
          # 行=都道府県 / 列=年度 の表を自動検出し、pref_data.json の年
          # (2014・2019 など) に一致する年度だけを取り込みます
          # 和暦の列見出し (平成26年度 / 令和元年度) にも対応
"""

import argparse
import json
import os
import re
import sys

from fetch_flows import pref_from_name, PREF_NAMES

BASE = os.path.dirname(os.path.abspath(__file__))
FLOW_OUT = os.path.join(BASE, "..", "data", "flow_data.json")
PREF_OUT = os.path.join(BASE, "..", "data", "pref_data.json")


def load_sheets(path, sheet=None):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl が必要です: pip install openpyxl")
    if path.lower().endswith(".xls"):
        sys.exit(".xls 形式は未対応です。Excel で .xlsx に保存し直してください。")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    names = [sheet] if sheet else wb.sheetnames
    for name in names:
        ws = wb[name]
        grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        yield name, grid


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace(" ", "").strip()
    if s in ("", "-", "…", "***", "X", "―"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_year(v):
    """「2019」「2019年度」「平成26年度」「令和元年度」などを西暦4桁に変換。"""
    s = str(v or "").strip()
    m = re.search(r"(19|20)\d{2}", s)
    if m:
        return m.group(0)
    for era, base in (("令和", 2018), ("平成", 1988), ("昭和", 1925)):
        m = re.search(era + r"\s*(元|\d+)", s)
        if m:
            n = 1 if m.group(1) == "元" else int(m.group(1))
            return str(base + n)
    return None


def find_pref_line(grid, axis):
    """都道府県名が40件以上並ぶ行 (axis='row') または列 (axis='col') を探す。
    戻り値: (行/列の位置, {位置: 都道府県コード})"""
    n_rows = len(grid)
    n_cols = max((len(r) for r in grid), default=0)
    rng = range(min(n_rows, 80)) if axis == "row" else range(min(n_cols, 40))
    best = None
    for i in rng:
        mapping = {}
        cells = (grid[i] if axis == "row"
                 else [(grid[r][i] if i < len(grid[r]) else None) for r in range(n_rows)])
        for j, cell in enumerate(cells):
            code = pref_from_name(cell)
            if code and code not in mapping.values():
                mapping[j] = code
        if len(mapping) >= 40 and (best is None or len(mapping) > len(best[1])):
            best = (i, mapping)
    return best if best else (None, None)


# ---------------------------------------------------------------- matrix (物流)
def cmd_matrix(args):
    for name, grid in load_sheets(args.file, args.sheet):
        header_row, col_map = find_pref_line(grid, "row")
        label_col, row_map = find_pref_line(grid, "col")
        if header_row is None or label_col is None:
            continue
        flows = {}
        for r, from_code in row_map.items():
            if r <= header_row:
                continue
            for c, to_code in col_map.items():
                if c == label_col or from_code == to_code:
                    continue
                v = to_num(grid[r][c] if c < len(grid[r]) else None)
                if v is None or v <= 0:
                    continue
                f, t = (to_code, from_code) if args.swap else (from_code, to_code)
                flows[(f, t)] = flows.get((f, t), 0) + v * args.scale
        if len(flows) < 500:
            continue
        print(f"シート「{name}」から {len(flows)} ペアを検出")

        out = {"meta": {"modes": {}}, "flows": {}}
        if os.path.exists(FLOW_OUT):
            with open(FLOW_OUT, encoding="utf-8") as fp:
                out = json.load(fp)
        out["meta"].pop("note", None)
        out["meta"].setdefault("modes", {})["goods"] = {
            "label": "貨物の流動 (貨物地域流動調査)", "unit": args.unit,
            "in": "流入", "out": "流出"}
        out.setdefault("flows", {})
        if not args.merge:
            out["flows"]["goods"] = {}   # 既定は物流を作り直す (--merge で追記)
        out["flows"].setdefault("goods", {})[str(args.year)] = \
            [[f, t, round(v)] for (f, t), v in flows.items() if v >= 1]
        with open(FLOW_OUT, "w", encoding="utf-8") as fp:
            json.dump(out, fp, ensure_ascii=False)
        smp = sorted(flows.items(), key=lambda x: -x[1])[:3]
        print(f"→ {FLOW_OUT} に {args.year} 年度として書き出しました")
        print("上位フロー (向きの確認用。逆に見える場合は --swap で再実行):")
        for (f, t), v in smp:
            print(f"  {PREF_NAMES[f]} → {PREF_NAMES[t]} : {round(v):,} {args.unit}")
        return
    sys.exit("47×47 の都道府県マトリクスを検出できませんでした。--sheet でシート名を"
             "指定するか、表の構造をご確認ください。")


# ---------------------------------------------------------------- kenmin (域際収支)
def sheet_pref(name, grid):
    """シート名または冒頭セルから都道府県を推定する。"""
    code = pref_from_name(name)
    if code:
        return code
    for r in range(min(6, len(grid))):
        for cell in grid[r][:6]:
            code = pref_from_name(cell)
            if code:
                return code
    return None


def find_item_year_layout(grid):
    """行=項目 / 列=年度 のレイアウトから「移出入」行と年度列を探す。
    戻り値: (item_row, {year: col}) or (None, None)"""
    item_row, best = None, None
    for r, row in enumerate(grid[:200]):
        for cell in row[:4]:
            n = str(cell or "")
            if "移出入" in n:
                key = (0 if "純" in n else 1)
                if best is None or key < best[0]:
                    best = (key, r)
    if best is None:
        return None, None
    item_row = best[1]
    year_cols = {}
    for r in range(min(item_row, 40)):
        cand = {}
        for c, cell in enumerate(grid[r]):
            y = parse_year(cell)
            if y and y not in cand:
                cand[y] = c
        if len(cand) >= 3:
            year_cols = cand
    return (item_row, year_cols) if year_cols else (None, None)


def write_netexp(pref_data, collected, targets, args):
    """collected = {year: {pref: 兆円}} を pref_data に書き込む。"""
    use = sorted(y for y in collected
                 if y in targets and len(collected[y]) >= 40)
    if not use:
        return False
    hit = 0
    for y in use:
        for code, v in collected[y].items():
            pref_data["prefs"][str(code)]["years"].setdefault(y, {})["netexp"] = v
            hit += 1
    pref_data["meta"].setdefault("metrics", {})["netexp"] = {
        "label": "域際収支 (財貨・サービスの移出入・純)",
        "short": "域際収支", "unit": "兆円/年度", "digits": 2}
    note = pref_data["meta"].get("note", "")
    if "県民経済計算" not in note:
        pref_data["meta"]["note"] = note + "。域際収支は県民経済計算 (年度)"
    with open(PREF_OUT, "w", encoding="utf-8") as fp:
        json.dump(pref_data, fp, ensure_ascii=False, indent=1)
    missing = sorted(targets - set(use))
    print(f"→ {hit} 件を取り込み (年度: {use})")
    if missing:
        print(f"※ Excel に無かった年: {missing} (地図ではグレー表示)")
    smp = sorted(collected[use[-1]].items(), key=lambda x: -x[1])[:3]
    print("上位 (妥当性確認用):",
          " / ".join(f"{PREF_NAMES[c]} {v:+.2f}兆円" for c, v in smp))
    return True


def cmd_kenmin(args):
    if not os.path.exists(PREF_OUT):
        sys.exit(f"{PREF_OUT} がありません。先に quickstart_japan.py を実行してください。")
    with open(PREF_OUT, encoding="utf-8") as fp:
        pref_data = json.load(fp)
    targets = set(args.years or pref_data.get("meta", {}).get("years", []))
    print(f"取り込み対象の年度: {sorted(targets)}")

    # ---- レイアウトB: 都道府県ごとのシート (行=項目, 列=年度) ----
    collected = {}
    matched_sheets = 0
    for name, grid in load_sheets(args.file, args.sheet):
        code = sheet_pref(name, grid)
        if code is None:
            continue
        item_row, year_cols = find_item_year_layout(grid)
        if item_row is None:
            continue
        matched_sheets += 1
        for y, c in year_cols.items():
            v = to_num(grid[item_row][c] if c < len(grid[item_row]) else None)
            if v is not None:
                collected.setdefault(y, {})[code] = round(v * args.scale, 3)
    if matched_sheets:
        print(f"都道府県別シート形式: {matched_sheets} シートから「移出入」を検出")
        if write_netexp(pref_data, collected, targets, args):
            return
        print("⚠ 対象年度と一致するデータが40都道府県分そろいませんでした。")
        print(f"  検出できた年度: {sorted(collected)[:10]}")

    # ---- レイアウトC: 年度ごとのシート (行=都道府県, 列=項目) ----
    collected = {}
    matched = []
    for name, grid in load_sheets(args.file, args.sheet):
        y = parse_year(name)
        if y is None or y not in targets:
            continue
        label_col, row_map = find_pref_line(grid, "col")
        if label_col is None:
            continue
        first_pref_row = min(row_map)
        # 都道府県行より上のヘッダから「移出入」を含む列を探す (「純」を優先)
        best = None
        for r in range(first_pref_row):
            for c, cell in enumerate(grid[r]):
                n = str(cell or "")
                if "移出入" in n:
                    key = (0 if "純" in n else 1)
                    if best is None or key < best[0]:
                        best = (key, c, n)
        if best is None:
            continue
        item_col, item_name = best[1], best[2]
        matched.append((name, item_name))
        for r, code in row_map.items():
            v = to_num(grid[r][item_col] if item_col < len(grid[r]) else None)
            if v is not None:
                collected.setdefault(y, {})[code] = round(v * args.scale, 3)
    if matched:
        print("年度別シート形式で検出:")
        for nm, it in matched:
            print(f"  シート「{nm}」 列「{it[:40]}」")
        if write_netexp(pref_data, collected, targets, args):
            return
        print("⚠ 40都道府県分そろいませんでした。--sheet の指定をご検討ください。")

    # ---- レイアウトA: 1シートに 行=都道府県 / 列=年度 ----
    for name, grid in load_sheets(args.file, args.sheet):
        label_col, row_map = find_pref_line(grid, "col")
        if label_col is None:
            continue
        # 年度の列を探す (都道府県行より上のどこかの行に年度見出しがある想定)
        year_cols = {}
        first_pref_row = min(row_map)
        for r in range(first_pref_row):
            for c, cell in enumerate(grid[r]):
                y = parse_year(cell)
                if y and c != label_col and c not in year_cols.values():
                    year_cols.setdefault(y, c)
        use = sorted(set(year_cols) & targets)
        if not use:
            continue
        print(f"シート「{name}」: 年度列 {sorted(year_cols)[:8]}{'...' if len(year_cols)>8 else ''}")

        collected2 = {}
        for y in use:
            c = year_cols[y]
            for r, code in row_map.items():
                v = to_num(grid[r][c] if c < len(grid[r]) else None)
                if v is not None:
                    collected2.setdefault(y, {})[code] = round(v * args.scale, 3)
        if write_netexp(pref_data, collected2, targets, args):
            return
    sys.exit("都道府県×年度の表を検出できませんでした。対象年度を含む Excel か、"
             "--sheet の指定をご確認ください。")


def main():
    p = argparse.ArgumentParser(description="公式 Excel → 消費アトラス用データ変換")
    sub = p.add_subparsers(dest="cmd", required=True)

    m = sub.add_parser("matrix", help="貨物地域流動調査の府県相互間マトリクスを変換")
    m.add_argument("file")
    m.add_argument("--year", required=True, help="年度 (例: 2023)")
    m.add_argument("--sheet", help="シート名 (未指定なら自動検出)")
    m.add_argument("--swap", action="store_true", help="発着の向きを反転")
    m.add_argument("--scale", type=float, default=1.0, help="値の換算係数")
    m.add_argument("--unit", default="千トン/年度", help="表示単位 (Excelの単位表記を確認)")
    m.add_argument("--merge", action="store_true",
                   help="既存の物流データに年を追記する (既定は作り直し)")
    m.set_defaults(func=cmd_matrix)

    k = sub.add_parser("kenmin", help="県民経済計算の移出入(純)を pref_data.json に統合")
    k.add_argument("file")
    k.add_argument("--years", nargs="*", help="取り込む年度 (未指定なら地図の年に合わせる)")
    k.add_argument("--sheet", help="シート名 (未指定なら自動検出)")
    k.add_argument("--scale", type=float, default=1e-6, help="兆円への換算 (既定: 百万円→兆円)")
    k.set_defaults(func=cmd_kenmin)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
